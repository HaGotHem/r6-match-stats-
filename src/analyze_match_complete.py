import json
import os
import sys
import argparse
from collections import defaultdict
import pandas as pd
from datetime import datetime

# Fix encoding for Windows console
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'replace')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'replace')

# Parse command line arguments
parser = argparse.ArgumentParser(description='Analyse des matchs R6 Siege')
parser.add_argument('--output-name', type=str, help='Nom du fichier de sortie (ex: Villa_2024-01-29_Ranked_2024-01-29-16-45.xlsx)')
parser.add_argument('--stats', type=str, default='kills,kost,survival,headshots,opening,multikills,plants,rating',
                    help='Stats a inclure (separes par virgule): kills,kost,survival,headshots,opening,multikills,plants,teamkills,rating')
args = parser.parse_args()

# Parser les stats demandees
ENABLED_STATS = set(args.stats.split(',')) if args.stats else set(['kills', 'kost', 'survival', 'headshots', 'opening', 'multikills', 'plants', 'rating'])

# Charger tous les rounds
rounds_data = []
MAX_ROUNDS = 20  # Support jusqu'à 20 rounds (overtime inclus)

for i in range(1, MAX_ROUNDS + 1):
    filename = f"data/match_data/round{i:02d}.json"
    if os.path.exists(filename) and os.path.getsize(filename) > 0:
        try:
            with open(filename, 'r', encoding='utf-8') as f:
                data = json.load(f)
                # Validation basique de la structure JSON
                if isinstance(data, dict) and 'players' in data and 'teams' in data:
                    rounds_data.append(data)
                    print(f"[OK] Round {i:02d} charge")
                else:
                    print(f"[WARNING] Round {i:02d} structure JSON invalide, ignore")
        except json.JSONDecodeError as e:
            print(f"[WARNING] Round {i:02d} JSON invalide ({e}), ignore")
        except Exception as e:
            print(f"[WARNING] Round {i:02d} erreur de lecture ({e}), ignore")
    elif i > 13:
        # Arrêter la recherche après 13 rounds si aucun fichier trouvé
        break
    else:
        print(f"[WARNING] Round {i:02d} vide ou manquant, ignore")

print(f"\nTotal rounds charges: {len(rounds_data)}")

# Structure pour stats par side (ATK / DEF / GLOBAL)
player_stats = defaultdict(lambda: {
    'atk': defaultdict(int),
    'def': defaultdict(int),
    'global': defaultdict(int),
    'team': None,
    'atk_kost_rounds': [],  # Liste des rounds où KOST est rempli en ATK
    'def_kost_rounds': [],  # Liste des rounds où KOST est rempli en DEF
})

ROUND_DURATION = 180  # secondes

# Parcourir tous les rounds
for round_num, round_data in enumerate(rounds_data, 1):
    # Déterminer qui est ATK et qui est DEF
    team_roles = {}
    for team in round_data['teams']:
        team_roles[team['name']] = team['role']  # "Attack" ou "Defense"

    # Mapper chaque joueur à son rôle dans ce round
    player_roles = {}
    for player in round_data['players']:
        username = player['username']
        team_index = player['teamIndex']
        team_name = round_data['teams'][team_index]['name']
        role = team_roles[team_name]  # "Attack" ou "Defense"
        player_roles[username] = role

        # Stocker l'équipe du joueur
        if player_stats[username]['team'] is None:
            player_stats[username]['team'] = team_index

    # Collecter les kills par joueur pour détecter les multi-kills
    round_kills_by_player = defaultdict(list)

    # Analyser les events du round
    for event in round_data.get('matchFeedback', []):
        event_type = event['type']['name']

        if event_type == 'Kill':
            killer = event['username']
            victim = event['target']
            headshot = event['headshot']
            time_sec = event['timeInSeconds']

            # Enregistrer le kill pour multi-kill detection
            round_kills_by_player[killer].append(time_sec)

            # Déterminer le side du killer
            if killer in player_roles:
                side = 'atk' if player_roles[killer] == 'Attack' else 'def'

                # Incrémenter kills
                player_stats[killer][side]['kills'] += 1
                player_stats[killer]['global']['kills'] += 1

                # Headshots
                if headshot:
                    player_stats[killer][side]['headshots'] += 1
                    player_stats[killer]['global']['headshots'] += 1

            # Vérifier si c'est un teamkill
            if killer in player_roles and victim in player_roles:
                if player_roles[killer] == player_roles[victim]:
                    side = 'atk' if player_roles[killer] == 'Attack' else 'def'
                    player_stats[killer][side]['teamkills'] += 1
                    player_stats[killer]['global']['teamkills'] += 1

        elif event_type == 'DefuserPlantComplete':
            planter = event['username']
            if planter in player_roles:
                side = 'atk' if player_roles[planter] == 'Attack' else 'def'
                player_stats[planter][side]['plants'] += 1
                player_stats[planter]['global']['plants'] += 1

        elif event_type == 'DefuserDisableComplete':
            disabler = event['username']
            if disabler in player_roles:
                side = 'atk' if player_roles[disabler] == 'Attack' else 'def'
                player_stats[disabler][side]['diffuses'] += 1
                player_stats[disabler]['global']['diffuses'] += 1

    # Détecter les multi-kills (2+ kills en 10 secondes)
    for killer, kill_times in round_kills_by_player.items():
        if len(kill_times) >= 2:
            # Trier par ordre croissant pour faciliter la comparaison
            kill_times_sorted = sorted(kill_times)

            # Vérifier si au moins 2 kills sont dans une fenêtre de 10 secondes
            has_multikill = False
            for i in range(len(kill_times_sorted) - 1):
                # Calculer la différence avec le kill suivant
                time_diff = kill_times_sorted[i + 1] - kill_times_sorted[i]
                if time_diff <= 10:
                    has_multikill = True
                    break

            if has_multikill and killer in player_roles:
                side = 'atk' if player_roles[killer] == 'Attack' else 'def'
                player_stats[killer][side]['multikills'] += 1
                player_stats[killer]['global']['multikills'] += 1

    # Opening kills/deaths (premier kill du round)
    if round_data.get('matchFeedback'):
        for event in round_data['matchFeedback']:
            if event.get('type', {}).get('name') == 'Kill':
                opener = event.get('username')
                victim = event.get('target')

                if opener and opener in player_roles:
                    side = 'atk' if player_roles[opener] == 'Attack' else 'def'
                    player_stats[opener][side]['opening_kills'] += 1
                    player_stats[opener]['global']['opening_kills'] += 1

                if victim and victim in player_roles:
                    side = 'atk' if player_roles[victim] == 'Attack' else 'def'
                    player_stats[victim][side]['opening_deaths'] += 1
                    player_stats[victim]['global']['opening_deaths'] += 1

                break  # Seulement le premier kill

    # Calculer temps de mort par joueur
    death_times = {}
    for event in round_data.get('matchFeedback', []):
        if event['type']['name'] == 'Kill':
            victim = event['target']
            time_of_death = event['timeInSeconds']
            death_times[victim] = time_of_death

    # Stats individuelles du round
    for stat in round_data.get('stats', []):
        username = stat['username']
        if username not in player_roles:
            continue

        side = 'atk' if player_roles[username] == 'Attack' else 'def'

        # Rounds joués
        player_stats[username][side]['rounds'] += 1
        player_stats[username]['global']['rounds'] += 1

        # Deaths
        if stat['died']:
            player_stats[username][side]['deaths'] += 1
            player_stats[username]['global']['deaths'] += 1

            # Temps de survie
            if username in death_times:
                survival_time = ROUND_DURATION - death_times[username]
                player_stats[username][side]['survival_time'] += survival_time
                player_stats[username]['global']['survival_time'] += survival_time
        else:
            # A survécu
            player_stats[username][side]['rounds_survived'] += 1
            player_stats[username]['global']['rounds_survived'] += 1
            player_stats[username][side]['survival_time'] += ROUND_DURATION
            player_stats[username]['global']['survival_time'] += ROUND_DURATION

        # KOST: Kill, Objective, Survived, or Traded
        kost_fulfilled = False
        if stat['kills'] > 0:  # Kill
            kost_fulfilled = True
        elif not stat['died']:  # Survived
            kost_fulfilled = True
        # Note: Objective et Traded nécessitent plus d'analyse, simplifié ici

        if kost_fulfilled:
            player_stats[username][f'{side}_kost_rounds'].append(round_num)

# Calculer les métriques finales
results = []
for username, stats in player_stats.items():
    row = {'Joueur': username}

    for side_key, side_label in [('atk', 'ATK'), ('def', 'DEF'), ('global', 'GLOBAL')]:
        s = stats[side_key]

        # Calculs de base
        kills = s.get('kills', 0)
        deaths = s.get('deaths', 0)
        rounds = s.get('rounds', 0)
        headshots = s.get('headshots', 0)
        rounds_survived = s.get('rounds_survived', 0)
        survival_time = s.get('survival_time', 0)
        opening_kills = s.get('opening_kills', 0)
        opening_deaths = s.get('opening_deaths', 0)

        # K/D
        kd = kills / max(deaths, 1)

        # +/-
        plus_minus = kills - deaths

        # KPR
        kpr = kills / max(rounds, 1)

        # HS%
        hs_pct = (headshots / max(kills, 1)) * 100 if kills > 0 else 0

        # Survival rate
        survival_rate = (rounds_survived / max(rounds, 1)) * 100 if rounds > 0 else 0

        # KOST
        kost_rounds_key = f'{side_key}_kost_rounds'
        kost_rounds_count = len(stats.get(kost_rounds_key, []))
        kost_pct = (kost_rounds_count / max(rounds, 1)) * 100 if rounds > 0 else 0

        # Temps de survie moyen
        avg_survival = survival_time / max(rounds, 1) if rounds > 0 else 0

        # Ratio opening
        opening_ratio = opening_kills / max(opening_deaths, 1) if opening_deaths > 0 else opening_kills

        # Rating (formule simplifiée)
        rating = (kpr * 0.7 + survival_rate/100 * 0.3) * 100 if rounds > 0 else 0

        # Ajouter colonnes selon les stats activees
        # Rating toujours en premier si active
        if 'rating' in ENABLED_STATS:
            row[f'{side_label} Rating'] = round(rating, 1)

        # Kills/Deaths/K-D
        if 'kills' in ENABLED_STATS:
            row[f'{side_label} Kills'] = kills
            row[f'{side_label} Deaths'] = deaths
            row[f'{side_label} K/D'] = round(kd, 2)
            row[f'{side_label} +/-'] = plus_minus
            row[f'{side_label} Rounds'] = rounds
            row[f'{side_label} KPR'] = round(kpr, 2)

        # KOST
        if 'kost' in ENABLED_STATS:
            row[f'{side_label} KOST%'] = round(kost_pct, 1)
            row[f'{side_label} Rounds KOST'] = kost_rounds_count

        # Survie
        if 'survival' in ENABLED_STATS:
            row[f'{side_label} Rounds Survie'] = rounds_survived
            row[f'{side_label} Temps Vie'] = f"{int(survival_time//60)}m{int(survival_time%60):02d}s"
            row[f'{side_label} Temps Moy (s)'] = round(avg_survival, 1)

        # Headshots
        if 'headshots' in ENABLED_STATS:
            row[f'{side_label} HS'] = headshots
            row[f'{side_label} HS%'] = round(hs_pct, 1)

        # Opening
        if 'opening' in ENABLED_STATS:
            row[f'{side_label} Opening K'] = opening_kills
            row[f'{side_label} Opening D'] = opening_deaths
            row[f'{side_label} Opening Ratio'] = round(opening_ratio, 2)

        # Multi-kills
        if 'multikills' in ENABLED_STATS:
            row[f'{side_label} Multi-kills'] = s.get('multikills', 0)

        # Plantes/Defuses
        if 'plants' in ENABLED_STATS:
            row[f'{side_label} Plantes'] = s.get('plants', 0)
            row[f'{side_label} Diffuses'] = s.get('diffuses', 0)

        # Teamkills
        if 'teamkills' in ENABLED_STATS:
            row[f'{side_label} Teamkills'] = s.get('teamkills', 0)

    row['Equipe'] = 'VOTRE EQUIPE' if stats['team'] == 0 else 'EQUIPE ENNEMIE'
    results.append(row)

# Créer DataFrame
df = pd.DataFrame(results)

# Trier par équipe puis par Rating global
df = df.sort_values(['Equipe', 'GLOBAL Rating'], ascending=[True, False])

# Reorganiser colonnes selon les stats activees
cols = ['Joueur', 'Equipe']
for side in ['ATK', 'DEF', 'GLOBAL']:
    if 'rating' in ENABLED_STATS:
        cols.append(f'{side} Rating')
    if 'kills' in ENABLED_STATS:
        cols.extend([
            f'{side} Kills',
            f'{side} Deaths',
            f'{side} K/D',
            f'{side} +/-',
            f'{side} Rounds',
            f'{side} KPR',
        ])
    if 'kost' in ENABLED_STATS:
        cols.extend([
            f'{side} KOST%',
            f'{side} Rounds KOST',
        ])
    if 'survival' in ENABLED_STATS:
        cols.extend([
            f'{side} Rounds Survie',
            f'{side} Temps Vie',
            f'{side} Temps Moy (s)',
        ])
    if 'headshots' in ENABLED_STATS:
        cols.extend([
            f'{side} HS',
            f'{side} HS%',
        ])
    if 'opening' in ENABLED_STATS:
        cols.extend([
            f'{side} Opening K',
            f'{side} Opening D',
            f'{side} Opening Ratio',
        ])
    if 'multikills' in ENABLED_STATS:
        cols.append(f'{side} Multi-kills')
    if 'plants' in ENABLED_STATS:
        cols.extend([
            f'{side} Plantes',
            f'{side} Diffuses',
        ])
    if 'teamkills' in ENABLED_STATS:
        cols.append(f'{side} Teamkills')

# Filtrer les colonnes qui existent dans le DataFrame
cols = [c for c in cols if c in df.columns]
df = df[cols]

# Sauvegarder avec mise en forme
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Generer le nom de fichier
if args.output_name:
    # Utiliser le nom fourni en parametre
    output_file = args.output_name
else:
    # Nom par defaut avec date et heure
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_file = f'Match_Stats_Complete_{timestamp}.xlsx'

def format_worksheet(ws, df, is_team_sheet=False):
    """Applique la mise en forme à une feuille Excel"""

    # Couleurs
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Bleu marine
    atk_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # Rouge clair
    def_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")     # Vert clair
    global_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Jaune clair
    team0_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")   # Bleu très clair
    team1_fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")   # Rouge très clair

    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Formater les en-têtes
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

        # Couleur de fond selon le type de colonne
        if 'ATK' in col_name:
            cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        elif 'DEF' in col_name:
            cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        elif 'GLOBAL' in col_name:
            cell.fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")

    # Formater les données
    for row_idx in range(2, ws.max_row + 1):
        # Alterner les couleurs par équipe
        equipe_cell = ws.cell(row=row_idx, column=2)  # Colonne Equipe
        row_fill = team0_fill if equipe_cell.value == 'VOTRE EQUIPE' else team1_fill

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Appliquer la couleur de ligne
            if col_idx <= 2:  # Joueur et Equipe
                cell.fill = row_fill
                if col_idx == 1:  # Nom du joueur en gras
                    cell.font = Font(bold=True, size=10)

    # Ajuster automatiquement la largeur des colonnes
    for col_idx, col_name in enumerate(df.columns, 1):
        column_letter = get_column_letter(col_idx)

        # Largeur par défaut selon le type de colonne
        if col_name == 'Joueur':
            width = 20
        elif col_name == 'Equipe':
            width = 18
        elif 'Temps Vie' in col_name:
            width = 12
        elif 'Rating' in col_name or 'KOST' in col_name:
            width = 10
        else:
            width = 9

        ws.column_dimensions[column_letter].width = width

    # Figer la première ligne
    ws.freeze_panes = 'A2'

    # Hauteur de ligne pour l'en-tête
    ws.row_dimensions[1].height = 30

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    # Stats complètes
    df.to_excel(writer, sheet_name='Stats Completes', index=False)
    format_worksheet(writer.sheets['Stats Completes'], df)

    # Feuille par équipe
    df_team0 = df[df['Equipe'] == 'VOTRE EQUIPE'].copy()
    df_team1 = df[df['Equipe'] == 'EQUIPE ENNEMIE'].copy()

    if not df_team0.empty:
        df_team0.to_excel(writer, sheet_name='VOTRE EQUIPE', index=False)
        format_worksheet(writer.sheets['VOTRE EQUIPE'], df_team0, is_team_sheet=True)

    if not df_team1.empty:
        df_team1.to_excel(writer, sheet_name='EQUIPE ENNEMIE', index=False)
        format_worksheet(writer.sheets['EQUIPE ENNEMIE'], df_team1, is_team_sheet=True)

print(f"\n{'='*60}")
print(f"[SUCCESS] Rapport complet genere: {output_file}")
print(f"{'='*60}")
print(f"\n{len(df)} joueurs analyses")
print(f"Colonnes par side: ATK / DEF / GLOBAL")
print(f"Total de {len(cols)-2} statistiques par joueur")
